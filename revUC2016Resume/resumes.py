import openpyxl;
#load two workbooks
checkin = openpyxl.load_workbook('sheet1.xlsx')
resume = openpyxl.load_workbook('sheet2.xlsx')
wb = openpyxl.Workbook()
ws = wb.active
# get the first sheet of each workbook
checkin = checkin[checkin.get_sheet_names()[0]]
resume = resume[resume.get_sheet_names()[0]]

checkinNames = list()
resumeNames = list()
#Column 4 of checkin is first names, column 5 is last names. Join for full names
for i in range(2, 223):
    name = checkin.cell(row = i, column = 4).internal_value + " " + checkin.cell(row = i, column = 5).internal_value
    name = name.lower()
    checkinNames.append(name)

for i in range(1, 300):
    name = resume.cell(row=i, column=1).internal_value
    name = name.lower()
    resumeNames.append(name)

attendedLinks = list()
noAttendedLinks = list()
for i in range(len(resumeNames)):
    link = resume.cell(row=i + 1, column=2).internal_value
    if resumeNames[i] in checkinNames:
        attendedLinks.append(link)
    else:
        noAttendedLinks.append(link)

for i in range(len(attendedLinks)):
    ws['A'+str(i+1)] = attendedLinks[i]

for i in range(len(noAttendedLinks)):
    ws['I'+str(i+1)] = noAttendedLinks[i]

wb.save('result.xlsx')
